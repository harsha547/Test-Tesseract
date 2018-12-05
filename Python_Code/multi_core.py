from PIL import Image
from wand.image import Image as PDFImg
import pytesseract
import cv2
import os
import shutil
from openpyxl import Workbook, load_workbook
from metadata import data
import time
from datetime import timedelta, datetime
import argparse
import sys
import logging
import multiprocessing
import re

PDF_BASE_DIRECTORY = '/home/ubuntu/Documents/pdfs'
IMAGES_DIRECTORY = '/home/ubuntu/Documents/Images'
CROP_DIRECTORY = '/home/ubuntu/Documents/Crop'
CROPPED_IMAGE_PATH = '/home/ubuntu/Documents/Crop/cropped.jpg'

STARTING_ROW = 2

HOUSE_HEADER_ITEMS = ['मकान न.:', 'मकान ने.:', 'समकान न:',
                      'समकान नं.:', 'मकान नें. :', 'स ?-', 'मकान नें.:',
                      'समकान न. : ', 'मकान नें.:', 'स', ' ?-', '?', '.','_',' _']

NAME_HEADER_ITEMS = ['.', '-', 'पति :', 'पिता :', 'पिता -','पत्ति :']


def replace_all(text, items):
    for i in items:
        text = text.replace(i, '')
    return text


def create_workbook():
    wb_excel = Workbook()
    wb_excel.save(EXCEL_BASE_PATH)


def delete_existing_images(folder_path):
        shutil.rmtree(folder_path)
        os.makedirs(folder_path)


def count_of_files(folder_path):
    return len([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))])


def crop_image(image_path, coords, saved_location):
    image_obj = Image.open(image_path)
    cropped_image = image_obj.crop(coords)
    cropped_image.save(saved_location)


def strip_newlines(text):
    text_list = text.split('\n')
    sent_str = ''
    for i in text_list:
        if i != '':
            sent_str += str(i) + '\n'
    text_list = sent_str.split('\n')
    new_str = []
    for count, i in enumerate(text_list):
        if count < 3:
            new_str.append(i)
        else:
            s = ' '
            new_str.append(s.join(text_list[3:]))
            return new_str
    return new_str


def update_excel(returned_data):
    global STARTING_ROW
    for image_data in returned_data:
        print(image_data)
        for record in image_data:
            sh_electors['A' + str(STARTING_ROW)] = record.get('ac_code', 'NA')
            sh_electors['B' + str(STARTING_ROW)] = record.get('ac_name', 'NA')
            sh_electors['C' + str(STARTING_ROW)] = record.get('ps_name', 'NA')
            sh_electors['D' + str(STARTING_ROW)] = record.get('anubagg_name', 'NA')
            sh_electors['E' + str(STARTING_ROW)] = record.get('Name', 'NA')
            sh_electors['F' + str(STARTING_ROW)] = record.get('Husband_Father_Name', 'NA')
            sh_electors['G' + str(STARTING_ROW)] = record.get('House Number', 'NA')
            sh_electors['H' + str(STARTING_ROW)] = record.get('Age', 'NA')
            sh_electors['I' + str(STARTING_ROW)] = record.get('Gender', 'NA')
            sh_electors['J' + str(STARTING_ROW)] = record.get('page_number', 'NA')
            sh_electors['K' + str(STARTING_ROW)] = record.get('position', 'NA')
            STARTING_ROW = STARTING_ROW + 1


def parse_text(text, position, page_number, anubagg_name, ac_data, ps_data):
    split_text = strip_newlines(text)
    record_data = {}
    for count, item in enumerate(split_text):
        if count == 0:
            record_data["Name"] = item.replace('नाम : ', '').strip()
        elif count == 1:
            if item.find('पति') != -1:
                record_data["Husband_Father_Name"] = replace_all(item.strip(), NAME_HEADER_ITEMS).strip()
            else:
                record_data["Husband_Father_Name"] = replace_all(item.strip(), NAME_HEADER_ITEMS).strip()
        elif count == 2:
            try:
                if item.find(':') != -1:
                    item = item.split(':')[1]
                elif item.find('-') != -1:
                    item = item.split('-')[1]
                elif item.find('..') != -1:
                    item = item.split('..')[1]
                else:
                    item = item.split('.')[1]
                record_data["House Number"] = replace_all(item, HOUSE_HEADER_ITEMS).strip()
                if str(record_data["House Number"]).strip() == "":
                    record_data["House Number"] = "NA"
            except:
                pass
        elif count == 3:
            try:
                age = str(int(re.search(r'\d+', item).group()))
                record_data["Age"] = str(age[-2:]).strip()
                if record_data["Age"] == "":
                    record_data["Age"] = "NA"
            except:
                pass
            try:
                if item.find('पुरुष') != -1:
                    record_data["Gender"] = 'पुरुष'
                else:
                    record_data["Gender"] = 'महिला'
                if str(record_data["Gender"]).strip() == "":
                    record_data["Gender"] = "NA"
            except:
                pass
        else:
            print("something wrong")
    record_data["page_number"] = str(page_number)
    record_data["position"] = str(position)
    try:
        record_data["ac_code"] = ac_data[0].strip()
        record_data["ac_name"] = ac_data[1].strip()
    except:
        record_data["ac_code"] = ""
        record_data["ac_name"] = ""
    try:
        record_data["ps_name"] = ps_data[1].strip()
    except:
        record_data["ps_name"] = ""
    try:
        record_data["anubagg_name"] = anubagg_name[1].strip()
    except:
        record_data["anubagg_name"] = ""
    return record_data


def parse_image(image_path, coords, crop_path):
    try:
        crop_image(image_path, coords, crop_path)
        image = cv2.imread(crop_path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.medianBlur(gray, 3)
        filename = "{}.png".format(os.getpid())
        cv2.imwrite(filename, gray)
        text = pytesseract.image_to_string(Image.open(filename), lang='hin2+hin')
    except Exception as e:
        return e
    return text


def parse_ac(image_path):
    crop_path = CROP_DIRECTORY + '/crop_AC.jpg'
    ac = parse_image(image_path, data.AC_Coordinates, crop_path)
    try:
        return ac.split("-")
    except Exception as e:
        return e


def parse_ps(image_path):
    crop_path = CROP_DIRECTORY + '/crop_PS.jpg'
    ps = parse_image(image_path, data.PS_Cooridnates, crop_path)
    try:
        return ps.split(":")
    except Exception as e:
        return e


def parse_anubhagg(image_path):
    crop_path = CROP_DIRECTORY + '/crop_Anubhagg.jpg'
    anubagg = parse_image(image_path, data.Anubagg_Coordinates, crop_path)
    try:
        return anubagg.split(":")
    except Exception as e:
        return e


def process_each_image(attr):
    returned_text = []
    image_path = attr['image_path']
    file_number = attr['file_number']
    page_number = attr['page_number']
    ac_data = parse_ac(image_path)
    ps_data = parse_ps(image_path)
    anubagg_name = parse_anubhagg(image_path)
    for i in range(1, 11):
        for j in range(1, 4):
                position = str(i) + ',' + str(j)
                coords = data.Coordinates[int(str(i)+str(j))]
                crop_path = CROP_DIRECTORY + '/crop_' + str(file_number) + '_' + str(i) + '_' + str(j) + '.jpg'
                crop_image(image_path, coords, crop_path)
                image = cv2.imread(crop_path)
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                gray = cv2.medianBlur(gray, 3)
                filename = "{}.png".format(os.getpid())
                cv2.imwrite(filename, gray)
                text = pytesseract.image_to_string(Image.open(filename), lang='hin2+hin')
                print('pageNumber : ' + str(page_number) + ' position : (' + str(i) + ',' + str(j) + ')')
                print(text)
                print('\n')
                returned_text.append(parse_text(text, position, page_number, anubagg_name, ac_data, ps_data))
    return returned_text


def process_images():
    files_count = count_of_files(IMAGES_DIRECTORY)
    list_of_images = []
    pool = multiprocessing.Pool(multiprocessing.cpu_count()-1)
    for file_number in range(2, files_count - 1):
        dict_meta = {}
        image_path = IMAGES_DIRECTORY + '/pdf_image-' + str(file_number) + '.jpg'
        dict_meta['image_path'] = image_path
        dict_meta['file_number'] = file_number
        dict_meta['page_number'] = file_number + 1
        list_of_images.append(dict_meta)
    results = pool.map(process_each_image, list_of_images)
    pool.close()
    pool.join()
    update_excel(results)
    delete_existing_images(CROP_DIRECTORY)


def convert_pdf_to_images(pdf_name):
    with PDFImg(filename=pdf_name, resolution=300) as img:
        img.compression_quality = 100
        img.save(filename=IMAGES_DIRECTORY + '/pdf_image.jpg')


def main(pdf_name):
    delete_existing_images(IMAGES_DIRECTORY)
    convert_pdf_to_images(pdf_name)
    process_images()


if __name__ == '__main__':
    global sh_electors
    global sh_metadata
    global EXCEL_BASE_PATH
    global LOG_FILE_PATH
    EXCEL_BASE_PATH = '/home/ubuntu/Documents/Excel_Files'
    LOG_FILE_PATH = '/home/ubuntu/Documents/Python_Code/'
    ap = argparse.ArgumentParser()
    ap.add_argument("-dst", "--DST_Folder", required=True, type=str, help="District Folder Name")
    ap.add_argument("-ac", "--AC_Folder", required=True, type=str, help="Assembly Consistutency Folder Name")
    args = vars(ap.parse_args())
    if args["DST_Folder"] is None or args["AC_Folder"] is None:
        print("Please mention District Folder and AC_Folder")
        sys.exit()
    files_directory = PDF_BASE_DIRECTORY + "/" + str(args["DST_Folder"]) + "/" + str(args["AC_Folder"])
    if os.path.exists(files_directory):
        print("Folder Exist")
    else:
        print("Folder Doesn't Exist, Please recheck naming convention")
        sys.exit()
    EXCEL_BASE_PATH = EXCEL_BASE_PATH + "/" + str(args["AC_Folder"]) + ".xlsx"
    LOG_FILE_PATH = LOG_FILE_PATH + "/" + str(args["AC_Folder"]) + ".log"
    start = time.time()
    create_workbook()
    excel_file = load_workbook(EXCEL_BASE_PATH)
    excel_file.create_sheet('electors')
    excel_file.create_sheet('metadata')
    sh_electors = excel_file['electors']
    sh_metadata = excel_file['metadata']
    sheet1 = excel_file['Sheet']
    excel_file.remove(sheet1)
    sh_electors['A1'].value = 'AC Code'
    sh_electors['B1'].value = 'AC Name'
    sh_electors['C1'].value = 'PS'
    sh_electors['D1'].value = 'Anubagg'
    sh_electors['E1'].value = 'Name'
    sh_electors['F1'].value = 'Father Name/Husband Name'
    sh_electors['G1'].value = 'House Number'
    sh_electors['H1'].value = 'Age'
    sh_electors['I1'].value = 'Gender'
    sh_electors['J1'].value = 'Page Number'
    sh_electors['K1'].value = 'Position'
    LOG_FORMAT = '%(message)s'
    logging.basicConfig(filename=LOG_FILE_PATH, format=LOG_FORMAT, filemode='w')
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    counter = 1
    for subdir, dirs, files in os.walk(files_directory):
        for file in files:
            if file != '.DS_Store':
                Pdf_path = os.path.join(subdir, file)
                print('Pdf_Name: ' + str(Pdf_path).rsplit('/', 1)[1])
                print('\n')
                start_timestamp = datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
                logger.info('COUNTER :' + str(counter) + ' Now processing, ' + ' ' + str(Pdf_path).rsplit('/', 1)[1] + ' ' +  start_timestamp)
                main(Pdf_path)
                counter = counter + 1
                excel_file.save(EXCEL_BASE_PATH)
                end_timestamp = datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
                logger.info('-----completed :-' + ' ' + end_timestamp + ' ' + '------')
                logger.info('\n')
    end = time.time()
    minutes = str(timedelta(seconds=(end - start)))
    logger.info( 'Program has taken {}'.format(minutes) )